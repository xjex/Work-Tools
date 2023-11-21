import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def spacing_excel_for_gdrive(json_file, output_file):
    # Load JSON data from file
    with open(json_file, 'r') as file:
        json_data = json.load(file)
    # Convert JSON data to a DataFrame and reverse the order
    df = pd.json_normalize(json_data).iloc[::-1].reset_index(drop=True)
    '''
    #not reversing the order
    df = pd.json_normalize(json_data)
    '''
    # Number of empty rows to insert
    num_empty_rows = 7

    # Create a new DataFrame with multiple empty rows inserted
    empty_rows_df = pd.concat([df] + [pd.DataFrame(index=df.index + (i + 1) * 0.14) for i in range(num_empty_rows)], axis=0).sort_index(kind='merge')
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the Excel sheet
    for c_idx, col_name in enumerate(empty_rows_df.columns, start=1):
        # Write the column name in the first row
        ws.cell(row=1, column=c_idx, value=col_name)

        # Set the background color to cyan for specific cells based on condition
        for r_idx, value in enumerate(empty_rows_df[col_name], start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if not pd.isna(value):
                cell.fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')  # Cyan color for non-empty cells

    # Save the Excel workbook
    wb.save(output_file)
    print(f'Data written to {output_file}')

if __name__ == "__main__":
    clean_data = "../Output/Puppeteer_web_cleaned.json"
    output_drive_file = '../Output/Puppeteer_Google_Drive_Ready.xlsx'
    spacing_excel_for_gdrive(clean_data, output_drive_file)
    print(f"Conversion successful. Excel file saved at {output_drive_file}")