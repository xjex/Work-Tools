import json
import re
import pandas as pd
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import datetime

def spacing_excel_for_gdrive(json_file, output_file):
    # Load JSON data from file
    with open(json_file, 'r') as file:
        json_data = json.load(file)
    # Convert JSON data to a DataFrame and reverse the order
    df = pd.json_normalize(json_data).iloc[::-1].reset_index(drop=True)
    '''
    #not reversing the order
    df = pd.json_normalize(json_data)
    '''
    # Number of empty rows to insert
    num_empty_rows = 7

    # Create a new DataFrame with multiple empty rows inserted
    empty_rows_df = pd.concat([df] + [pd.DataFrame(index=df.index + (i + 1) * 0.14) for i in range(num_empty_rows)], axis=0).sort_index(kind='merge')
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write the DataFrame to the Excel sheet
    for c_idx, col_name in enumerate(empty_rows_df.columns, start=1):
        # Write the column name in the first row
        ws.cell(row=1, column=c_idx, value=col_name)

        # Set the background color to cyan for specific cells based on condition
        for r_idx, value in enumerate(empty_rows_df[col_name], start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if not pd.isna(value):
                cell.fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')  # Cyan color for non-empty cells

    # Save the Excel workbook
    wb.save(output_file)
    print(f'Data written to {output_file}')


def json_to_excel(json_file, excel_file):
    # Read JSON file into a Pandas DataFrame
    df = pd.read_json(json_file)
    df.to_excel(excel_file, index=False,)

def clean_json_file(input_file, output_file):
    with open(input_file, 'r') as file:
        data = json.load(file)

    cleaned_data = clean_json(data)

    with open(output_file, 'w') as file:
        json.dump(cleaned_data, file, indent=2)

def clean_json(data):
    cleaned_data = []
    for entry in data:
        contribution_text = entry.get("Contribution", "")
        cleaned_contribution = remove_time_related_strings(contribution_text)
        entry["Contribution"] = cleaned_contribution
        cleaned_data.append(entry)
    return cleaned_data

def remove_time_related_strings(contribution_text):
    time_patterns = [
        r"Contributed \d+ day[s]? ago by",
        r"Contributed about \d+ month[s]? ago by",
        r"Contributed \d+ month[s]? ago by ",
        r"Contributed about \d+ hours ago by",
        r"Contributed about \d+ hour[s]? ago by",
        r"Contributed \d+ hour[s]? ago by",
    ]

    time_regex = re.compile("|".join(time_patterns))
    cleaned_contribution = time_regex.sub("", contribution_text)
    return cleaned_contribution.strip()

if __name__ == "__main__":
    # Create a folder for the output files
    current_datetime = datetime.datetime.now()
    date = current_datetime.strftime("%Y-%m-%d")
    folder_path = f'../Output/stackdata/{date}'
    os.makedirs(folder_path, exist_ok=True)


    input_json_file = f'../Output/stackdata/{date}/Puppeteer_web_scraped_data.json'
    output_json_file = f'../Output/stackdata/{date}/Puppeteer_web_cleaned.json'

    os.system('cls' if os.name == 'nt' else 'clear')
    clean_json_file(input_json_file, output_json_file)
    print(f"Cleaning successful. Cleaned JSON file saved at {output_json_file}")

    
    clean_data = f'../Output/stackdata/{date}/Puppeteer_web_cleaned.json'
    output_excel_file = f'../Output/stackdata/{date}/Puppeteer_Clean_Excel.xlsx'
    
    json_to_excel(clean_data, output_excel_file)
    print(f"Conversion successful. Excel file saved at {output_excel_file}")

    output_drive_file = f'../Output/stackdata/{date}/Puppeteer_Google_Drive_Ready.xlsx'
    spacing_excel_for_gdrive(clean_data, output_drive_file)
    print(f"GoogleDrive Format. Excel file saved at {output_drive_file}")
